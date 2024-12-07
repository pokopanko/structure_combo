import os
import tkinter as tk
from tkinter import ttk
from typing import Dict, List, Optional

import openpyxl


def get_folder_structure(root_path: str) -> Dict[str, Dict]:
    """
    フォルダ構成を取得する関数。
    指定されたフォルダをルートに、そのサブフォルダを階層構造で取得する。

    Args:
        root_path (str): ルートフォルダのパス。

    Returns:
        dict: フォルダ構成をネストされた辞書で表現。
    """

    def build_structure(path: str) -> Dict[str, Dict]:
        structure = {}
        for entry in os.scandir(path):
            if entry.is_dir():
                structure[entry.name] = build_structure(entry.path)
        return structure

    if not os.path.exists(root_path):
        raise ValueError(f"指定されたパスが存在しません: {root_path}")
    return build_structure(root_path)


class FolderComboBox(tk.Frame):
    """
    フォルダ構成をインデント付きで表示するコンボボックス。
    """

    def __init__(
        self,
        master: tk.Widget,
        folder_structure: Dict[str, Dict],
        root_path: str,
        *args,
        **kwargs,
    ):
        super().__init__(master, *args, **kwargs)

        self.combobox: ttk.Combobox = ttk.Combobox(self, state="readonly")
        self.combobox.pack(fill=tk.X, expand=True)

        self.folder_structure: Dict[str, Dict] = folder_structure
        self.root_path: str = root_path
        self.populate_combobox()

    def populate_combobox(self) -> None:
        """
        フォルダ構造をインデント付きでコンボボックスにセットする。
        """

        def traverse_structure(
            structure: Dict[str, Dict], parent_path: str = ""
        ) -> List[str]:
            items = []
            for folder_name, sub_structure in structure.items():
                full_path = os.path.join(parent_path, folder_name)
                items.append(full_path)
                items.extend(traverse_structure(sub_structure, full_path))
            return items

        folder_items = traverse_structure(self.folder_structure, self.root_path)
        self.combobox["values"] = folder_items


class ExcelSheetSelector(tk.Frame):
    """
    フォルダ内のエクセルファイル名とシート名を表示するクラス。
    """

    def __init__(self, master: tk.Widget, *args, **kwargs):
        super().__init__(master, *args, **kwargs)

        self.folder_path: Optional[str] = None

        self.excel_file_label: tk.Label = tk.Label(self, text="Excel File:")
        self.excel_file_label.pack(anchor=tk.W, padx=10, pady=5)

        self.sheet_combobox: ttk.Combobox = ttk.Combobox(self, state="readonly")
        self.sheet_combobox.pack(fill=tk.X, padx=10, pady=5)

    def update_folder(self, folder_path: str) -> None:
        """
        選択されたフォルダを更新して、エクセルファイルとシート名を表示。
        """
        self.folder_path = folder_path.strip()
        self.check_and_display_excel_file()

    def check_and_display_excel_file(self) -> None:
        """
        フォルダ内のエクセルファイルを確認して、シート名をコンボボックスに表示。
        """
        if not self.folder_path:
            return

        excel_files = [
            f
            for f in os.listdir(self.folder_path)
            if f.endswith(".xlsx") and not f.startswith("~$")
        ]

        if len(excel_files) == 1:
            excel_file = excel_files[0]
            self.excel_file_label.config(text=f"Excel File: {excel_file}")

            # シート名を取得してコンボボックスにセット
            file_path = os.path.join(self.folder_path, excel_file)
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            self.sheet_combobox["values"] = workbook.sheetnames
            if workbook.sheetnames:
                self.sheet_combobox.current(0)
            workbook.close()
        else:
            self.excel_file_label.config(text="Excel File: None (or multiple files)")
            self.sheet_combobox["values"] = []
            self.sheet_combobox.set("")  # コンボボックスをクリア


class MainApp(tk.Tk):
    """
    全体の統合アプリケーション。
    """

    def __init__(self) -> None:
        super().__init__()
        self.title("Folder and Excel Viewer")
        self.geometry("500x300")

        # フォルダ構造取得
        root_folder: str = r"C:\\work\\output"  # 実際のルートフォルダを設定
        folder_structure: Dict[str, Dict] = get_folder_structure(root_folder)

        # フォルダコンボボックス
        folder_frame: tk.LabelFrame = tk.LabelFrame(self, text="Folder Structure")
        folder_frame.pack(fill=tk.X, padx=10, pady=10)

        self.folder_combobox: FolderComboBox = FolderComboBox(
            folder_frame, folder_structure, root_folder
        )
        self.folder_combobox.pack(fill=tk.X, expand=True, padx=10, pady=5)

        # エクセルシートセレクタ
        excel_frame: tk.LabelFrame = tk.LabelFrame(self, text="Excel File and Sheets")
        excel_frame.pack(fill=tk.X, padx=10, pady=10)

        self.excel_selector: ExcelSheetSelector = ExcelSheetSelector(excel_frame)
        self.excel_selector.pack(fill=tk.X, expand=True, padx=10, pady=5)

        # フォルダ選択時の処理
        self.folder_combobox.combobox.bind(
            "<<ComboboxSelected>>", self.on_folder_selected
        )

    def on_folder_selected(self, event: tk.Event) -> None:
        """
        フォルダ選択時に呼ばれる処理。
        """
        selected_folder: str = self.folder_combobox.combobox.get().strip()
        if os.path.isdir(selected_folder):
            self.excel_selector.update_folder(selected_folder)


# 実行（環境に応じてローカルで試してください）
MainApp().mainloop()
